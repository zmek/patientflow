import json
from io import BytesIO

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from typing import List, Dict


# Parsing the 'Duration in department' column to extract start and end times
def parse_duration(duration):
    # Check if the delimiter " to " is in the duration string
    if " to " in duration:
        parts = duration.split(" to ")
    # If not, check for the delimiter " - "
    elif " - " in duration:
        parts = duration.split(" - ")
    # Lastly, check for the delimiter "-  " (dash followed by two spaces)
    elif "-  " in duration:
        parts = duration.split("-  ")
    else:
        # If none of the delimiters are found, return default values
        return duration, None

    # If the split operation was successful and resulted in two parts, return them
    if len(parts) == 2:
        return parts[0], parts[1]
    else:
        # If there aren't exactly two parts, return default values
        return parts, None


def load_data_into_dataframe(url: str, sheet_name: str, columns: List[str]) -> pd.DataFrame:
    """
    Load data into a DataFrame from a specified URL and sheet name.

    Args:
        url (str): URL to the data file.
        sheet_name (str): Name of the sheet to load.
        columns (list): List of columns to load.

    Returns:
        pandas.DataFrame: Loaded data as a DataFrame.

    """
    # Specify the 'Duration in department' target string to find the duration column
    target_string = "Duration in department"

    response = requests.get(url)

    if response.status_code == 200:
        file_in_memory = BytesIO(response.content)

        if url.endswith(".xlsx"):
            workbook = load_workbook(file_in_memory, data_only=True)
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                print(f"The sheet '{sheet_name}' does not exist in the workbook.")
                return None

            header_row = None
            column_indices = []
            column_headers = []
            duration_col_index = None
            adjust_header_row = False  # Flag to adjust the header row if needed

            # Find the 'Duration in department' header row
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith(
                        target_string
                    ):
                        header_row = cell.row
                        duration_col_index = cell.column
                        break
                if header_row:
                    break

            if header_row is None:
                print("Header 'Duration in department' not found.")
                return None

            # Initially set 'Duration in department' as the first column
            column_indices.append(duration_col_index)
            column_headers.append("Duration in department")

            # Check the immediate row below for other column headers
            below_header_row_values = (
                [cell.value for cell in sheet[header_row + 1]]
                if header_row + 1 <= sheet.max_row
                else []
            )
            header_row_values = [cell.value for cell in sheet[header_row]]

            for col_name in columns:
                if col_name in below_header_row_values:
                    column_indices.append(below_header_row_values.index(col_name) + 1)
                    column_headers.append(col_name)
                    adjust_header_row = (
                        True  # Adjust if any column is found in the row below
                    )
                elif col_name in header_row_values:
                    column_indices.append(header_row_values.index(col_name) + 1)
                    column_headers.append(col_name)

            # Adjust the data extraction start if headers are in the next row
            data_start_row = header_row + 2 if adjust_header_row else header_row + 1

            # Extract data from selected rows and columns
            data = []
            for row in sheet.iter_rows(
                min_row=data_start_row, max_row=data_start_row + 25
            ):
                row_data = [row[idx - 1].value for idx in column_indices]
                data.append(row_data)

        elif url.endswith(".xls"):
            import xlrd

            workbook = xlrd.open_workbook(file_contents=file_in_memory.getvalue())
            try:
                sheet = workbook.sheet_by_name(sheet_name)
            except xlrd.biffh.XLRDError:
                print(f"The sheet '{sheet_name}' does not exist in the workbook.")
                return None
            # Dynamically determine the number of columns
            max_column = sheet.ncols
            column_headers = sheet.row_values(11, 0, max_column)
            data = [
                sheet.row_values(row_idx, 0, max_column) for row_idx in range(12, 38)
            ]
        else:
            print("Unsupported file format.")
            return None

        # Creating a DataFrame
        df = pd.DataFrame(data, columns=column_headers)

        # Remove last row if all values are NaN
        if df.iloc[-1].isnull().all():
            df = df.iloc[:-1]

        # Parse the duration columns
        df[["start_time", "end_time"]] = df["Duration in department"].apply(
            lambda x: pd.Series(parse_duration(x))
        )

        df["end_time"] = df["end_time"].apply(
            lambda x: (
                int(x.split(":")[0]) * 60 + int(x.split(":")[1])
                if x is not None
                else None
            )
        )

        return df
    else:
        print(f"Failed to download the Excel file. Status code: {response.status_code}")
        return None


def load_json_configuration(json_file_path: str, reference_year: str) -> Dict:
    """
    Load NHSE source data configuration from a JSON file.

    Args:
        json_file_path (str): Path to the JSON file.
        reference_year (str): Year for which data is to be analyzed.

    Returns:
        dict: A dictionary containing the URL, sheet name, and columns.

    """
    try:
        with open(json_file_path) as file:
            nhse_urls_dict = json.load(file)
        return nhse_urls_dict[reference_year]
    except (FileNotFoundError, KeyError) as e:
        raise ValueError("Error loading configuration: " + str(e))


def calculate_cumulative_proportions(df: pd.DataFrame, columns: List[str]) -> pd.DataFrame:
    """
    Calculate cumulative totals and proportions of patients.

    Args:
        df (pandas.DataFrame): The DataFrame containing patient data.
        columns (list): Columns to sum over.

    Returns:
        pandas.DataFrame: DataFrame with added cumulative totals and proportions.

    """
    columns_to_sum = set(columns).intersection(df.columns)
    total_patients = df[list(columns_to_sum)].sum().sum()
    df["cum_total"] = df[list(columns_to_sum)].sum(axis=1).cumsum()
    df["cum_prop"] = df["cum_total"] / total_patients
    return df


def interpolate_probabilities(df: pd.DataFrame, prediction_window: int, time_interval: int) -> np.ndarray:
    """
    Interpolate probabilities at specified time intervals.

    NB - this currently assumes there are 24 (6 per hour over 4 hours) 10 min slots in the first 24 rows of df

    Args:
        df (pandas.DataFrame): The DataFrame with time and cumulative proportion data.
        prediction_window (int): Total prediction window in hours.
        time_interval (int): Time interval in minutes.

    Returns:
        numpy.ndarray: Interpolated probabilities.

    """
    df_new = pd.concat(
        [
            df[["end_time", "cum_prop"]].iloc[:24],
            pd.DataFrame({"end_time": [prediction_window], "cum_prop": [1]}),
        ]
    )
    xnew = np.arange(1, prediction_window / time_interval + 1) * time_interval
    return np.interp(xnew, df_new["end_time"], df_new["cum_prop"])


def compute_rolling_mean(interpolated_probs: np.ndarray) -> np.ndarray:
    """
    Compute the rolling mean of interpolated probabilities.

    Args:
        interpolated_probs (numpy.ndarray): Interpolated probabilities.

    Returns:
        numpy.ndarray: Smoothed probabilities.

    """
    rolling_mean = np.convolve(interpolated_probs, np.ones(2) / 2, mode="valid")
    rolling_mean = np.insert(rolling_mean, 0, interpolated_probs[0])
    return np.flip(rolling_mean)


def calculate_probability(
    json_file_path: str, reference_year: str, prediction_window: int, time_interval: int
) -> np.ndarray:
    """
    Calculate the probability of hospital admission within a specified prediction window.

    Args:
        json_file_path (str): Path to the JSON file.
        reference_year (str): Year for which the data is to be analyzed.
        prediction_window (int): Total prediction window (in minutes).
        time_interval (int): Time interval (in minutes).

    Returns:
        numpy.ndarray: Array of probabilities.

    """
    # Validate and Load Configuration
    nhse_source_data = load_json_configuration(json_file_path, reference_year)

    # Validate and Load Data
    df = load_data_into_dataframe(
        nhse_source_data["url"],
        nhse_source_data["sheet_name"],
        nhse_source_data["columns"],
    )

    # Validate and Compute Cumulative Proportions
    df = calculate_cumulative_proportions(df, nhse_source_data["columns"])

    # Validate and Interpolate Probabilities
    interpolated_probs = interpolate_probabilities(df, prediction_window, time_interval)

    # Validate and Compute Rolling Mean
    return compute_rolling_mean(interpolated_probs)
